import os
import webbrowser
import openpyxl
from pathlib import Path

# load excel with its path
data_folder = "C:/Users/crich/Downloads/aha-music-export_2022-11-15.xlsx"
print(data_folder)
wrkbk = openpyxl.load_workbook(data_folder)

print(data_folder)
sheet = wrkbk.active
print(sheet)
  
# iterate through excel and display data
for i in range(1, sheet.max_row+1):
    url = sheet.cell(row=1, column=6).hyperlink.target
    webbrowser.open(url)
